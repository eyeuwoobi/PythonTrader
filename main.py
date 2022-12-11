#2022-11-26

"""
- 기업가치 = 자기자본 + (초과이익/할인율)
: 자기자본 = 지배기업소유주지분(지배주주지분)
: 할인율 = BBB- 등급 회사채 5년 수익률 (매일 갱신) https://www.kisrating.com/ratingsStatistics/statics_spread.do
: 초과이익 = 자기자본 * (가중평균 3년 ROE - 할인율)
- 적정주가 = 기업가치 / 유통주식수

"""

import requests
from bs4 import BeautifulSoup
import xlrd
import openpyxl

# 전역 변수 선언
## 할인율
discountRateLocation: int  = 87 #BBB- 5y
KISlink: str = 'https://www.kisrating.com/ratingsStatistics/statics_spread.do'

## 다트 재무제표
DART_API_Key = 'b64695f3f2a79d07bde772ffa630f935e0d050c0'



def getDiscountRate(link, location) -> float:
    html = requests.get(link).text
    soup = BeautifulSoup(html, "html.parser")
    result = float(soup.select('td.ar.pr40')[location].text)
    return result

a = getDiscountRate(KISlink, discountRateLocation)
print("Today's discount rate :", a)


'''
#res = requests.get('https://opendart.fss.or.kr/api/company.json', params={'crtfc_key' : api_key, 'corp_code' : samsung})
res = requests.get('https://www.fnspace.com/Api/FinanceApi?key=sample&format=json&code=A005930&item=M111000,M111100,M111600,M113000&consolgb=M&annualgb=A&fraccyear=2014&toaccyear=2018')
res.raise_for_status()
print(res)
data = res.json()
print(data)
'''



# Define variable to load the dataframe
dataframe = openpyxl.load_workbook("/workspace/PythonTrader/stocksDB.xlsx")

# Define variable to read sheet
dataframe1 = dataframe.active

# Iterate the loop to read the cell values

for row in dataframe1.iter_rows(0, dataframe1.max_row):
    print(row[1].value)